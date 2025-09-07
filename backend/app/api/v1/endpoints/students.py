from typing import List, Optional
from fastapi import APIRouter, Depends, HTTPException, status, Query, UploadFile, File, BackgroundTasks, Response
from fastapi.responses import Response
from sqlalchemy.orm import Session

from app.crud import crud_student
from app.schemas import student as student_schema
from app.services.generators import generate_roll_number, generate_institutional_email
from app.core.config import settings
import csv
import io
import smtplib
from email.mime.text import MIMEText
from sqlalchemy.exc import IntegrityError
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from app.db.session import get_db

router = APIRouter()

@router.post("/", response_model=student_schema.Student, status_code=status.HTTP_201_CREATED)
def create_student(
    *,
    db: Session = Depends(get_db),
    student_in: student_schema.StudentCreate,
):
    """
    Create new student.
    """
    # Check if email already exists
    student = crud_student.get_student_by_email(db, email=student_in.email)
    if student:
        raise HTTPException(
            status_code=400,
            detail="A student with this email already exists."
        )
    
    # Check if admission number already exists
    student = crud_student.get_student_by_admission_number(
        db, admission_number=student_in.admission_number
    )
    if student:
        raise HTTPException(
            status_code=400,
            detail="A student with this admission number already exists."
        )
    
    return crud_student.create_student(db=db, student=student_in)

@router.post("/import/preview")
async def import_preview(file: UploadFile = File(...)):
    if not (file.filename.endswith('.csv') or file.filename.endswith('.xlsx')):
        raise HTTPException(status_code=400, detail="Only CSV and Excel (.xlsx) files are supported")
    
    content = await file.read()
    
    if file.filename.endswith('.xlsx'):
        # Handle Excel file
        df = pd.read_excel(io.BytesIO(content))
        rows = df.to_dict('records')
    else:
        # Handle CSV file
        text = content.decode('utf-8')
        reader = csv.DictReader(io.StringIO(text))
        rows = list(reader)
    
    # Validate minimal columns present
    required = [
        'Name','Address','Gender','Category','Date of Birth','Phone Number','Branch','Year','Mother Name'
    ]
    preview = []
    for row in rows:
        errors = []
        for r in required:
            if not row.get(r) or pd.isna(row.get(r)):
                errors.append(f"Missing {r}")
        preview.append({
            'row': row,
            'is_valid': len(errors) == 0,
            'errors': errors,
        })
    return { 'total': len(rows), 'preview': preview }

@router.post("/import/save")
async def import_save(
    *,
    db: Session = Depends(get_db),
    file: UploadFile = File(...),
):
    if not (file.filename.endswith('.csv') or file.filename.endswith('.xlsx')):
        raise HTTPException(status_code=400, detail="Only CSV and Excel (.xlsx) files are supported")
    
    content = await file.read()
    
    if file.filename.endswith('.xlsx'):
        # Handle Excel file
        df = pd.read_excel(io.BytesIO(content))
        rows = df.to_dict('records')
    else:
        # Handle CSV file
        text = content.decode('utf-8')
        reader = csv.DictReader(io.StringIO(text))
        rows = list(reader)
    
    total = 0
    successful = 0
    failed = 0
    errors: List[str] = []
    for index, row in enumerate(rows, start=2):
        total += 1
        try:
            name = str(row.get('Name') or '').strip()
            if pd.isna(row.get('Name')):
                name = ''
            if ' ' in name:
                first_name, last_name = name.split(' ', 1)
            else:
                first_name, last_name = name, ''
            roll_number = generate_roll_number()
            dept = str(row.get('Branch') or 'dept')
            if pd.isna(row.get('Branch')):
                dept = 'dept'
            email_institute = generate_institutional_email(name, dept)
            # ensure unique institutional email
            if email_institute and '@' in email_institute:
                local_part, domain_part = email_institute.split('@', 1)
                unique_email = email_institute
                suffix = 1
                while crud_student.get_student_by_email(db, unique_email):
                    unique_email = f"{local_part}{suffix}@{domain_part}"
                    suffix += 1
                email_institute = unique_email

            # ensure unique admission/roll by regenerating if needed
            while crud_student.get_student_by_admission_number(db, roll_number):
                roll_number = generate_roll_number()
            # Use institutional email as the required email to avoid invalid cases like
            # "first.@example.com" when last_name is missing
            # Handle potential NaN values from Excel
            phone_val = row.get('Phone Number')
            if pd.isna(phone_val):
                phone_val = None
            else:
                phone_val = str(phone_val)
            
            dob_val = row.get('Date of Birth')
            if pd.isna(dob_val):
                dob_val = None
            else:
                dob_val = str(dob_val)
            
            gender_val = row.get('Gender')
            if pd.isna(gender_val):
                gender_val = 'male'
            else:
                gender_val = str(gender_val).lower()
            
            address_val = row.get('Address')
            if pd.isna(address_val):
                address_val = None
            else:
                address_val = str(address_val)
            
            year_val = row.get('Year')
            if pd.isna(year_val):
                year_val = None
            else:
                year_val = str(year_val)
            
            payload = student_schema.StudentCreate(
                first_name=first_name,
                last_name=last_name,
                email=email_institute,
                phone=phone_val,
                date_of_birth=dob_val,
                gender=student_schema.Gender[gender_val],
                address=address_val,
                state=year_val,
                country='India',
                postal_code=None,
                admission_number=roll_number,
                roll_number=roll_number,
                institutional_email=email_institute,
                department=dept,
            )
            # dedupe by email or admission_number before insert
            if crud_student.get_student_by_email(db, payload.email) or \
               crud_student.get_student_by_admission_number(db, payload.admission_number):
                failed += 1
                errors.append(f"Row {index}: Duplicate record (email or roll/admission already exists)")
            else:
                crud_student.create_student(db, payload)
                successful += 1
        except IntegrityError as e:
            db.rollback()
            failed += 1
            errors.append(f"Row {index}: {str(e)}")
        except Exception as e:
            failed += 1
            errors.append(f"Row {index}: {str(e)}")
    
    # Return proper JSON response
    from fastapi.responses import JSONResponse
    return JSONResponse(
        content={
            'total': total, 
            'successful': successful, 
            'failed': failed, 
            'errors': errors
        },
        status_code=200
    )

@router.get("/export/csv")
def export_students_csv(db: Session = Depends(get_db)):
    students = crud_student.get_students(db)
    headers = [
        'Roll Number','First Name','Last Name','Institutional Email','Personal Email','Phone','Address','Gender','Department','Year'
    ]
    output = io.StringIO()
    writer = csv.writer(output, quoting=csv.QUOTE_ALL)
    writer.writerow(headers)
    for s in students:
        writer.writerow([
            getattr(s, 'roll_number', '') or '',
            s.first_name or '',
            s.last_name or '',
            getattr(s, 'institutional_email', '') or '',
            s.email or '',
            s.phone or '',
            s.address or '',
            s.gender.value if s.gender else '',
            getattr(s, 'department', '') or '',
            getattr(s, 'state', '') or ''
        ])
    return Response(output.getvalue(), media_type="text/csv", headers={
        "Content-Disposition": "attachment; filename=students.csv"
    })

@router.get("/export/excel")
def export_students_excel(db: Session = Depends(get_db)):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        students = crud_student.get_students(db)
        
        # Create a new workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Students"
        
        # Define headers
        headers = [
            'Roll Number', 'First Name', 'Last Name', 'Institutional Email', 
            'Personal Email', 'Phone', 'Address', 'Gender', 'Department', 'Year'
        ]
        
        # Add headers to first row
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Add student data
        for row, student in enumerate(students, 2):
            ws.cell(row=row, column=1, value=getattr(student, 'roll_number', '') or '')
            ws.cell(row=row, column=2, value=student.first_name or '')
            ws.cell(row=row, column=3, value=student.last_name or '')
            ws.cell(row=row, column=4, value=getattr(student, 'institutional_email', '') or '')
            ws.cell(row=row, column=5, value=student.email or '')
            ws.cell(row=row, column=6, value=student.phone or '')
            ws.cell(row=row, column=7, value=student.address or '')
            ws.cell(row=row, column=8, value=student.gender.value if student.gender else '')
            ws.cell(row=row, column=9, value=getattr(student, 'department', '') or '')
            ws.cell(row=row, column=10, value=getattr(student, 'state', '') or '')
        
        # Style the header row
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Set column widths
        column_widths = [15, 20, 20, 35, 30, 15, 40, 10, 35, 12]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=students.xlsx"}
        )
        
    except Exception as e:
        print(f"Excel export error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to export Excel file: {str(e)}")

@router.post("/email/csv")
def email_students_csv(
    *,
    db: Session = Depends(get_db),
    background: BackgroundTasks,
    recipient: str,
):
    def _send(csv_content: str):
        if not settings.SMTP_USER or not settings.SMTP_PASSWORD or not settings.SMTP_FROM_EMAIL:
            return
        msg = MIMEText(csv_content)
        msg['Subject'] = 'Students Export'
        msg['From'] = settings.SMTP_FROM_EMAIL
        msg['To'] = recipient
        with smtplib.SMTP(settings.SMTP_HOST, settings.SMTP_PORT) as server:
            server.starttls()
            server.login(settings.SMTP_USER, settings.SMTP_PASSWORD)
            server.sendmail(settings.SMTP_FROM_EMAIL, [recipient], msg.as_string())

    csv_content = export_students_csv(db)
    background.add_task(_send, csv_content)
    return { 'status': 'scheduled' }

@router.get("/", response_model=List[student_schema.Student])
def read_students(
    db: Session = Depends(get_db),
    skip: int = 0,
    limit: int = 100,
    search: Optional[str] = Query(None, description="Search by name, email, or admission number"),
):
    """
    Retrieve students with optional search and pagination.
    """
    students = crud_student.get_students(
        db, skip=skip, limit=limit, search=search
    )
    return students

@router.get("/count", response_model=int)
def get_students_count(
    db: Session = Depends(get_db),
):
    """
    Get total count of students.
    """
    return crud_student.get_students_count(db)

@router.get("/{student_id}", response_model=student_schema.Student)
def read_student(
    student_id: int,
    db: Session = Depends(get_db),
):
    """
    Get student by ID.
    """
    student = crud_student.get_student(db, student_id=student_id)
    if not student:
        raise HTTPException(
            status_code=404,
            detail="Student not found"
        )
    return student

@router.put("/{student_id}", response_model=student_schema.Student)
def update_student(
    *,
    db: Session = Depends(get_db),
    student_id: int,
    student_in: student_schema.StudentUpdate,
):
    """
    Update a student.
    """
    student = crud_student.get_student(db, student_id=student_id)
    if not student:
        raise HTTPException(
            status_code=404,
            detail="Student not found"
        )
    
    # Check if email is being updated to an existing email
    if student_in.email and student_in.email != student.email:
        existing_student = crud_student.get_student_by_email(
            db, email=student_in.email
        )
        if existing_student:
            raise HTTPException(
                status_code=400,
                detail="A student with this email already exists."
            )
    
    # Check if admission number is being updated to an existing one
    if student_in.admission_number and student_in.admission_number != student.admission_number:
        existing_student = crud_student.get_student_by_admission_number(
            db, admission_number=student_in.admission_number
        )
        if existing_student:
            raise HTTPException(
                status_code=400,
                detail="A student with this admission number already exists."
            )
    
    return crud_student.update_student(
        db, db_student=student, student_in=student_in
    )

@router.delete("/{student_id}")
def delete_student(
    *,
    db: Session = Depends(get_db),
    student_id: int,
):
    """
    Delete a student.
    """
    student = crud_student.get_student(db, student_id=student_id)
    if not student:
        raise HTTPException(
            status_code=404,
            detail="Student not found"
        )
    return crud_student.delete_student(db=db, student_id=student_id)
